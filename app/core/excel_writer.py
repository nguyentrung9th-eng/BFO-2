import os
import re
import unicodedata
import openpyxl

def remove_vietnamese_accents(text: str) -> str:
    if not text:
        return ""
    # Chuyển ký tự unicode dạng tổ hợp thành dựng sẵn và loại bỏ dấu
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = text.replace('Đ', 'D').replace('đ', 'd')
    # Giữ lại các chữ cái, chữ số và khoảng trắng
    text = re.sub(r'[^a-zA-Z0-9\s]', '', text)
    return text.strip().upper()

def write_to_excel(template_path, output_path, extracted_data_list, global_defaults):
    """
    Đọc template (mẫu mới .xlsx), điền dữ liệu mới bắt đầu từ dòng 3 (dòng 2 là Header),
    hoặc hỗ trợ cả file .xls cũ.
    """
    try:
        is_xlsx = False
        field_col_map = {
            'STT': 1, 'Diễn Giải': 2, 'Mã hoạt động (NS)': 3, 'TT Chịu phí': 4,
            'Mã chi phí': 5, 'Số tiền (A)': 6, 'Thuế (B)': 7, 'Thành tiền (A+B)': 8,
            'MSNV': 9, 'ASM': 10, 'Hình Thức Thanh Toán': 11, 'KTPT': 12,
            'GHI CHÚ': 13, 'MST': 14, 'TÊN NCC': 15
        }

        # Nếu template_path rỗng hoặc không tồn tại, tự khởi tạo workbook mới từ đầu
        if not template_path or not os.path.exists(template_path):
            is_xlsx = True
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trich phi"
            # Tạo dòng 2 (Header)
            headers = [
                'STT', 'Diễn Giải', 'Mã hoạt động (NS)', 'TT Chịu phí', 'Mã chi phí',
                'Số tiền (A)', 'Thuế (B)', 'Thành tiền (A+B)', 'MSNV', 'ASM',
                'Hình Thức Thanh Toán', 'KTPT', 'GHI CHÚ', 'MST', 'TÊN NCC'
            ]
            for col_num, header_title in enumerate(headers, 1):
                ws.cell(row=2, column=col_num, value=header_title)
                
        # Nếu template là file .xlsx (mẫu mới)
        elif template_path.endswith('.xlsx'):
            is_xlsx = True
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

        if is_xlsx:
            # Xóa các dòng dữ liệu mẫu cũ từ dòng 3 đến hết
            if ws.max_row >= 3:
                ws.delete_rows(3, ws.max_row - 2)

            current_row = 3
            stt = 1
            last_thang_nam = "07.26"

            for data in extracted_data_list:
                if not data:
                    continue

                ngay = str(data.get('NgayChungTu', ''))
                thang_nam = "07/2026"
                if ngay and '/' in ngay:
                    parts = ngay.split('/')
                    if len(parts) >= 2:
                        thang_nam = f"{parts[1].zfill(2)}/{parts[2]}" if len(parts) == 3 else ngay
                        if len(parts) == 3:
                            last_thang_nam = f"{parts[1].zfill(2)}.{parts[2][-2:]}"

                so_hd = str(data.get('SoChungTuNgoai', ''))
                if '|' in so_hd:
                    so_hd = so_hd.split('|')[-1].strip()
                else:
                    so_hd = so_hd.strip()

                bv = data.get('DienGiai_BoSung', '').strip()
                bv_suffix = f"_{bv}" if bv else ""
                dien_giai = f"THANH TOÁN CHI PHÍ TIẾP KHÁCH THEO KẾ HOẠCH THÁNG {thang_nam} _{so_hd}{bv_suffix}"

                # Tiền
                pre_tax = data.get('SoTienTinhThueGTGT', '')
                tax_amt = data.get('SoTienThueVAT', '')
                post_tax = data.get('ThanhTien', '')

                ws.cell(row=current_row, column=field_col_map['STT'], value=stt)
                ws.cell(row=current_row, column=field_col_map['Diễn Giải'], value=dien_giai)
                ws.cell(row=current_row, column=field_col_map['Mã hoạt động (NS)'], value=global_defaults.get('MaHoatDong', 'LSF0200'))
                ws.cell(row=current_row, column=field_col_map['TT Chịu phí'], value=global_defaults.get('TTChiuPhi', 'LSF'))
                ws.cell(row=current_row, column=field_col_map['Mã chi phí'], value=global_defaults.get('MaChiPhi', '1512'))

                cell_f = ws.cell(row=current_row, column=field_col_map['Số tiền (A)'], value=int(pre_tax) if str(pre_tax).isdigit() else pre_tax)
                cell_f.number_format = '#,##0'
                
                cell_g = ws.cell(row=current_row, column=field_col_map['Thuế (B)'], value=int(tax_amt) if str(tax_amt).isdigit() else tax_amt)
                cell_g.number_format = '#,##0'
                
                # Cột H: Thành tiền (A+B) dùng công thức Excel hoặc value
                cell_h = ws.cell(row=current_row, column=field_col_map['Thành tiền (A+B)'], value=f"=F{current_row}+G{current_row}")
                cell_h.number_format = '#,##0'

                ws.cell(row=current_row, column=field_col_map['MSNV'], value=global_defaults.get('MSNV', '7649'))
                ws.cell(row=current_row, column=field_col_map['ASM'], value=global_defaults.get('ASM', 'Hồ Vĩnh Hữu'))
                ws.cell(row=current_row, column=field_col_map['Hình Thức Thanh Toán'], value=global_defaults.get('HinhThucThanhToan', 'TM'))
                ws.cell(row=current_row, column=field_col_map['KTPT'], value=global_defaults.get('KTPT', 'Thúy'))
                ws.cell(row=current_row, column=field_col_map['GHI CHÚ'], value=global_defaults.get('GhiChu', 'ISP'))
                
                cell_mst = ws.cell(row=current_row, column=field_col_map['MST'], value=str(data.get('MaSoThue', '')))
                cell_mst.number_format = '@'
                
                ws.cell(row=current_row, column=field_col_map['TÊN NCC'], value=data.get('NhaCungCap', ''))

                current_row += 1
                stt += 1

            # Ghi công thức tính tổng cho ô F1, G1, H1 từ hàng 3 đến hàng 15
            ws['F1'] = "=SUM(F3:F15)"
            ws['F1'].number_format = '#,##0'

            ws['G1'] = "=SUM(G3:G15)"
            ws['G1'].number_format = '#,##0'

            ws['H1'] = "=SUM(H3:H15)"
            ws['H1'].number_format = '#,##0'

            # Đổi tên Sheet theo cấu trúc: Trich phi {TÊN_NGƯỜI_THỰC_HIỆN_KHÔNG_DẤU_IN_HOA} T{mm.yy}
            raw_person_name = global_defaults.get('TenNguoiThucHien', '')
            name_no_accent = remove_vietnamese_accents(raw_person_name)
            if not name_no_accent:
                name_no_accent = "THUY"

            sheet_title = f"Trich phi {name_no_accent} T{last_thang_nam}"
            
            # Nếu vượt quá 31 ký tự, ưu tiên xóa dần từ đầu tiên trong Tên người thực hiện
            if len(sheet_title) > 31:
                name_words = name_no_accent.split()
                while len(name_words) > 1 and len(sheet_title) > 31:
                    name_words.pop(0)  # Xóa từ đầu tiên
                    shortened_name = " ".join(name_words)
                    sheet_title = f"Trich phi {shortened_name} T{last_thang_nam}"

            ws.title = sheet_title[:31]  # Đảm bảo tối đa 31 ký tự theo quy định Excel Sheet Title

            if not output_path.endswith('.xlsx'):
                output_path = output_path.replace('.xls', '.xlsx')
            
            out_dir = os.path.dirname(output_path)
            if out_dir and not os.path.exists(out_dir):
                os.makedirs(out_dir, exist_ok=True)

            wb.save(output_path)
            return {"success": True, "output_path": output_path}

        # Fallback hỗ trợ xlrd/xlwt cho .xls cũ
        import xlrd
        import xlwt
        from xlutils.copy import copy

        if not output_path.endswith('.xls'):
            output_path = output_path.replace('.xlsx', '.xls')
            
        rb = xlrd.open_workbook(template_path, formatting_info=True)
        sheet = rb.sheet_by_index(0)
        wb = copy(rb)
        ws = wb.get_sheet(0)
        
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.height = 8 * 20
        style.font = font

        current_row = 1
        for data in extracted_data_list:
            if not data: continue
            so_hd = str(data.get('SoChungTuNgoai', ''))
            ws.write(current_row, 0, so_hd, style)
            current_row += 1

        wb.save(output_path)
        return {"success": True, "output_path": output_path}
        
    except Exception as e:
        return {"success": False, "error": f"Lỗi ghi Excel: {str(e)}"}
